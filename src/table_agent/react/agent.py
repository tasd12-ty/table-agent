"""ReAct Agent - 多轮观察-思考-行动循环"""

from __future__ import annotations

import json
import logging
import re
import shutil
import tempfile
from pathlib import Path

from ..config import ReactConfig
from ..llm import LLMClient
from .executor import CodeExecutor
from .models import ExecutionResult, ReactResult, ReactStep
from .prompts import (
    REACT_SYSTEM_PROMPT,
    REACT_USER_FIRST_ROUND,
    REACT_USER_FOLLOW_UP,
    REACT_USER_TEXT_ONLY,
)
from .renderer import SpreadsheetRenderer
from .tracer import TraceRecorder

logger = logging.getLogger(__name__)


class ReactAgent:
    """基于 ReAct 模式的多轮电子表格处理 agent"""

    def __init__(
        self,
        llm: LLMClient,
        config: ReactConfig,
        renderer: SpreadsheetRenderer | None = None,
        executor: CodeExecutor | None = None,
        tracer: TraceRecorder | None = None,
    ):
        self.llm = llm
        self.config = config
        self.renderer = renderer or SpreadsheetRenderer(backend=config.renderer_backend)
        self.executor = executor or CodeExecutor(timeout=config.code_timeout)
        self.tracer = tracer

    async def run(
        self,
        instruction: str,
        input_xlsx: str,
        answer_sheet: str = "",
        data_position: str = "",
        model: str | None = None,
        task_id: str = "",
    ) -> ReactResult:
        """执行 ReAct 循环

        Args:
            instruction: 自然语言任务指令
            input_xlsx: 输入 xlsx 文件路径
            answer_sheet: 答案所在 sheet 名称
            data_position: 数据位置描述
            model: LLM 模型覆盖
            task_id: 任务 ID

        Returns:
            ReactResult 包含完整执行轨迹
        """
        used_model = model or self.llm.default_model

        # 创建临时工作目录，保留原始扩展名
        work_dir = tempfile.mkdtemp(prefix="react_")
        ext = Path(input_xlsx).suffix or ".xlsx"
        work_filename = f"input{ext}"
        work_xlsx = Path(work_dir) / work_filename
        shutil.copy2(input_xlsx, work_xlsx)

        # 预分析输入文件
        file_analysis = self._analyze_input(str(work_xlsx))
        vba_hint = self._extract_vba_hint(input_xlsx)

        steps: list[ReactStep] = []
        messages: list[dict] = [
            {"role": "system", "content": REACT_SYSTEM_PROMPT}
        ]

        try:
            for round_num in range(1, self.config.max_rounds + 1):
                logger.info("任务 %s: 第 %d/%d 轮", task_id, round_num, self.config.max_rounds)

                # 1. 观察：渲染电子表格截图
                screenshots = await self.renderer.render(str(work_xlsx))

                # 保存截图和 xlsx 快照
                screenshot_paths: list[str] = []
                spreadsheet_path = ""
                if self.tracer:
                    screenshot_paths = self.tracer.save_screenshots(task_id, round_num, screenshots)
                    spreadsheet_path = self.tracer.save_spreadsheet(task_id, round_num, str(work_xlsx))

                # 2. 构建 user message
                if round_num == 1:
                    text = REACT_USER_FIRST_ROUND.format(
                        instruction=instruction,
                        filename=work_filename,
                        answer_sheet=answer_sheet,
                        data_position=data_position,
                        file_analysis=file_analysis,
                        vba_hint=vba_hint,
                        round=round_num,
                        max_rounds=self.config.max_rounds,
                    )
                else:
                    last_step = steps[-1]
                    exec_result = last_step.execution_result
                    if exec_result:
                        feedback = f"返回码: {exec_result.return_code}\n"
                        if exec_result.stdout:
                            feedback += f"标准输出:\n{exec_result.stdout[:2000]}\n"
                        if exec_result.stderr:
                            feedback += f"错误输出:\n{exec_result.stderr[:2000]}\n"
                        if exec_result.timed_out:
                            feedback += "（代码执行超时）\n"
                    else:
                        feedback = "（无执行结果）"

                    text = REACT_USER_FOLLOW_UP.format(
                        execution_feedback=feedback,
                        round=round_num,
                        max_rounds=self.config.max_rounds,
                    )

                # 如果没有截图，附加文本形式的表格内容
                if not screenshots:
                    spreadsheet_text = self._read_xlsx_as_text(str(work_xlsx))
                    text = REACT_USER_TEXT_ONLY.format(spreadsheet_text=spreadsheet_text) + text

                # 构建多模态 content blocks
                content: list[dict] = [{"type": "text", "text": text}]
                for img_b64 in screenshots:
                    content.append({
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{img_b64}"},
                    })

                # 上下文管理：仅保留最近 2 轮的截图，旧轮次替换为纯文本
                messages = self._manage_context(messages, round_num)
                messages.append({"role": "user", "content": content})

                # 3. 调用 LLM
                response = await self.llm.chat(
                    messages=messages,
                    model=used_model,
                )

                # 记录 assistant 回复
                messages.append({"role": "assistant", "content": response})

                # 4. 解析响应
                thought, action, code = self._parse_response(response)

                step = ReactStep(
                    round=round_num,
                    thought=thought,
                    action=action,
                    code=code,
                    raw_response=response,
                    screenshot_paths=screenshot_paths,
                    spreadsheet_path=spreadsheet_path,
                )

                # 5. 执行
                if action == "done":
                    steps.append(step)
                    logger.info("任务 %s: agent 宣布完成", task_id)
                    break

                if action == "code" and code:
                    exec_result = await self.executor.execute(code, work_dir)
                    step.execution_result = exec_result
                    steps.append(step)

                    if exec_result.success:
                        logger.info("任务 %s: 第 %d 轮代码执行成功", task_id, round_num)
                    else:
                        logger.warning("任务 %s: 第 %d 轮代码执行失败", task_id, round_num)
                else:
                    steps.append(step)
                    logger.warning("任务 %s: 第 %d 轮无有效代码", task_id, round_num)

            result = ReactResult(
                task_id=task_id,
                instruction=instruction,
                input_file=input_xlsx,
                output_file=str(work_xlsx) if work_xlsx.exists() else None,
                steps=steps,
                total_rounds=len(steps),
                success=any(s.action == "done" for s in steps),
                model_used=used_model,
            )

            # 保存完整轨迹
            if self.tracer:
                self.tracer.save_trace(result)

            return result

        except Exception as e:
            logger.error("任务 %s 执行异常: %s", task_id, e)
            return ReactResult(
                task_id=task_id,
                instruction=instruction,
                input_file=input_xlsx,
                output_file=None,
                steps=steps,
                total_rounds=len(steps),
                success=False,
                error_message=str(e),
                model_used=used_model,
            )

    @staticmethod
    def _parse_response(response: str) -> tuple[str, str, str | None]:
        """解析 LLM 的 JSON 响应

        Returns:
            (thought, action, code)
        """
        # 尝试直接解析 JSON
        try:
            data = json.loads(response)
            return (
                data.get("thought", ""),
                data.get("action", "code"),
                data.get("code"),
            )
        except json.JSONDecodeError:
            pass

        # 尝试提取 ```json ... ``` 代码块
        match = re.search(r"```json\s*\n(.*?)\n```", response, re.DOTALL)
        if match:
            try:
                data = json.loads(match.group(1))
                return (
                    data.get("thought", ""),
                    data.get("action", "code"),
                    data.get("code"),
                )
            except json.JSONDecodeError:
                pass

        # 回退：将整个响应作为 thought，不执行代码
        logger.warning("无法解析 LLM 响应为 JSON，跳过本轮")
        return (response[:500], "done", None)

    @staticmethod
    def _manage_context(messages: list[dict], current_round: int) -> list[dict]:
        """管理上下文：移除旧轮次的截图以控制 token 用量

        保留最近 2 轮的截图，更早的轮次仅保留文本。
        """
        if current_round <= 2:
            return messages

        managed = []
        for msg in messages:
            if msg["role"] == "user" and isinstance(msg.get("content"), list):
                # 检查是否是较早轮次的消息（粗略判断：前面的 user 消息）
                user_count = sum(1 for m in managed if m["role"] == "user")
                # 保留最近 2 轮的图片
                if user_count < current_round - 3:
                    # 移除图片，只保留文本
                    text_parts = [
                        p for p in msg["content"]
                        if p.get("type") == "text"
                    ]
                    managed.append({
                        "role": "user",
                        "content": text_parts if text_parts else msg["content"],
                    })
                else:
                    managed.append(msg)
            else:
                managed.append(msg)

        return managed

    @staticmethod
    def _analyze_input(xlsx_path: str) -> str:
        """预分析输入文件结构，供第一轮 prompt 使用"""
        import openpyxl

        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            parts: list[str] = []
            for ws in wb.worksheets:
                parts.append(f"- Sheet '{ws.title}': {ws.max_row} 行 × {ws.max_column} 列")
                # 读取前 2 行作为表头参考
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = [str(c) if c is not None else "" for c in row]
                    parts.append(f"  表头: {' | '.join(headers)}")
                    break
            wb.close()
            return "\n".join(parts) if parts else "（无法解析文件结构）"
        except Exception as e:
            return f"（文件分析失败: {e}）"

    @staticmethod
    def _extract_vba_hint(xlsx_path: str) -> str:
        """检测 .xlsm 文件中的 VBA 模块，返回提示信息"""
        if not xlsx_path.endswith(".xlsm"):
            return ""
        try:
            from zipfile import ZipFile
            with ZipFile(xlsx_path) as z:
                vba_files = [n for n in z.namelist() if "vbaProject" in n]
                if vba_files:
                    return "⚠️ 注意：输入文件包含 VBA 宏模块。请根据任务指令理解宏的意图，并用 Python/openpyxl 实现等效逻辑。\n"
        except Exception:
            pass
        return ""

    @staticmethod
    def _read_xlsx_as_text(xlsx_path: str, max_rows: int = 50) -> str:
        """将 xlsx 读取为文本表格（LibreOffice 不可用时的回退）"""
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        parts: list[str] = []

        for ws in wb.worksheets:
            parts.append(f"### Sheet: {ws.title}")
            rows_shown = 0
            for row in ws.iter_rows(values_only=True):
                if rows_shown >= max_rows:
                    parts.append(f"... (省略剩余行，共 {ws.max_row} 行)")
                    break
                cells = [str(c) if c is not None else "" for c in row]
                parts.append(" | ".join(cells))
                rows_shown += 1
            parts.append("")

        wb.close()
        return "\n".join(parts)
