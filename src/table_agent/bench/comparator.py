"""单元格对比引擎 - 比较 agent 输出与标准答案"""

from __future__ import annotations

import logging
import math
import re

import openpyxl

from .models import CellRange, ComparisonResult

logger = logging.getLogger(__name__)


class SpreadsheetComparator:
    """对比输出 xlsx 与答案 xlsx 在指定位置的单元格值"""

    def compare(
        self,
        output_path: str,
        answer_path: str,
        answer_position: str,
        answer_sheet: str,
    ) -> ComparisonResult:
        """对比输出与答案

        Args:
            output_path: agent 生成的 xlsx 路径
            answer_path: 标准答案 xlsx 路径
            answer_position: 答案位置描述 (如 "A3:D32", "'Sheet1'!A1:F5")
            answer_sheet: 默认答案 sheet 名

        Returns:
            ComparisonResult
        """
        try:
            output_wb = openpyxl.load_workbook(output_path, data_only=True)
            answer_wb = openpyxl.load_workbook(answer_path, data_only=True)
        except Exception as e:
            return ComparisonResult(error=f"无法打开文件: {e}")

        try:
            ranges = self._parse_answer_position(answer_position, answer_sheet)
            total_cells = 0
            matched_cells = 0
            mismatches: list[dict] = []

            for cell_range in ranges:
                sheet_name = cell_range.sheet

                # 获取对应 sheet
                if sheet_name not in output_wb.sheetnames:
                    mismatches.append({
                        "sheet": sheet_name,
                        "error": f"输出文件中缺少 sheet: {sheet_name}",
                    })
                    continue

                if sheet_name not in answer_wb.sheetnames:
                    logger.warning("答案文件中缺少 sheet: %s", sheet_name)
                    continue

                output_ws = output_wb[sheet_name]
                answer_ws = answer_wb[sheet_name]

                t, m, mm = self._compare_range(output_ws, answer_ws, cell_range)
                total_cells += t
                matched_cells += m
                mismatches.extend(mm)

            accuracy = matched_cells / total_cells if total_cells > 0 else 0.0

            return ComparisonResult(
                total_cells=total_cells,
                matched_cells=matched_cells,
                accuracy=accuracy,
                mismatches=mismatches[:50],  # 限制 mismatch 数量
            )

        except Exception as e:
            logger.error("对比出错: %s", e)
            return ComparisonResult(error=str(e))

        finally:
            output_wb.close()
            answer_wb.close()

    @staticmethod
    def _parse_answer_position(position: str, default_sheet: str) -> list[CellRange]:
        """解析 answer_position 字符串

        处理格式:
        - "A3:D32" (简单范围)
        - "K12" (单个单元格)
        - "'Sheet1'!A1:F5" (带 sheet 名的范围)
        - "'OUT CAS'!A2:C1529,'OUT CAS'!E2:G586" (多范围)
        - "'MINUS'!B2:E11,'PLUS'!B2:E5200" (多 sheet 多范围)
        """
        ranges: list[CellRange] = []

        # 按逗号分割，但要处理引号内的逗号
        segments = SpreadsheetComparator._split_ranges(position)

        for seg in segments:
            seg = seg.strip()
            if not seg:
                continue

            if "!" in seg:
                # 带 sheet 名: 'Sheet Name'!A1:B5 or Sheet1!A1:B5
                sheet_part, cell_part = seg.rsplit("!", 1)
                sheet_name = sheet_part.strip().strip("'")
            else:
                sheet_name = default_sheet
                cell_part = seg

            if ":" in cell_part:
                start, end = cell_part.split(":", 1)
                ranges.append(CellRange(sheet=sheet_name, start_cell=start, end_cell=end))
            else:
                ranges.append(CellRange(sheet=sheet_name, start_cell=cell_part))

        return ranges

    @staticmethod
    def _split_ranges(position: str) -> list[str]:
        """按逗号分割范围，处理引号内的逗号"""
        segments: list[str] = []
        current = ""
        in_quotes = False

        for ch in position:
            if ch == "'":
                in_quotes = not in_quotes
                current += ch
            elif ch == "," and not in_quotes:
                segments.append(current)
                current = ""
            else:
                current += ch

        if current:
            segments.append(current)

        return segments

    @staticmethod
    def _compare_range(
        output_ws,
        answer_ws,
        cell_range: CellRange,
    ) -> tuple[int, int, list[dict]]:
        """对比指定范围内的单元格

        Returns:
            (total_cells, matched_cells, mismatches)
        """
        start = cell_range.start_cell
        end = cell_range.end_cell or cell_range.start_cell

        total = 0
        matched = 0
        mismatches: list[dict] = []

        try:
            answer_cells = answer_ws[f"{start}:{end}"]
            output_cells = output_ws[f"{start}:{end}"]
        except Exception as e:
            return 0, 0, [{"error": f"无法读取范围 {start}:{end}: {e}"}]

        # 确保是元组的元组（即使单行/单列）
        if not isinstance(answer_cells, tuple):
            answer_cells = ((answer_cells,),)
        if not isinstance(output_cells, tuple):
            output_cells = ((output_cells,),)

        for a_row, o_row in zip(answer_cells, output_cells):
            if not isinstance(a_row, tuple):
                a_row = (a_row,)
            if not isinstance(o_row, tuple):
                o_row = (o_row,)

            for a_cell, o_cell in zip(a_row, o_row):
                total += 1
                a_val = a_cell.value
                o_val = o_cell.value

                if SpreadsheetComparator._cells_equal(o_val, a_val):
                    matched += 1
                else:
                    mismatches.append({
                        "cell": a_cell.coordinate,
                        "sheet": cell_range.sheet,
                        "expected": str(a_val),
                        "actual": str(o_val),
                    })

        return total, matched, mismatches

    @staticmethod
    def _cells_equal(output_val, answer_val) -> bool:
        """对比两个单元格值

        - None/空: 视为相等
        - 数值: 容差 1e-6 相对误差
        - 字符串: strip + 大小写不敏感
        """
        # 都是空
        if output_val is None and answer_val is None:
            return True
        if output_val == "" and answer_val is None:
            return True
        if output_val is None and answer_val == "":
            return True

        # 都是数值
        try:
            o_num = float(output_val)
            a_num = float(answer_val)
            if a_num == 0:
                return abs(o_num) < 1e-9
            return abs(o_num - a_num) / max(abs(a_num), 1e-9) < 1e-6
        except (TypeError, ValueError):
            pass

        # 字符串比较
        o_str = str(output_val).strip() if output_val is not None else ""
        a_str = str(answer_val).strip() if answer_val is not None else ""
        return o_str.lower() == a_str.lower()
